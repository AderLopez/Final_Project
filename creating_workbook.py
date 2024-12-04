import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def Creating_workbook(data,mean,median,std):
    try:
        #creating a virtual workbook:
        #This information and chart will be in sheet 1:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet2 = wb.create_sheet("Statistics")
        sheet3 = wb.create_sheet("Bar Chart")
        sheet4 = wb.create_sheet("Line Chart")
        sheet5 = wb.create_sheet("Pie Chart")
        sheet6 = wb.create_sheet("Scatter Plot")


        #We create our data/ your data can get from API
        print(data)

        #We need to add the headers first:
        Column_titles = data[0].keys()
        sheet.append(list(Column_titles))

        #Excel is taking the data as string, forcing the converting to int:
        for item in data:
            item['Rank'] = int(item['Rank'])

        #Excel is taking the data as string, forcing the converting to float:
        for item in data:
            item['priceUsd'] = float(item['priceUsd'])

        #Adding the data to the file:
        for item in data:
            sheet.append(list(item.values()))


        #Sheet 2 will be for storing the statistics data:

        sheet2['A1'] = 'Mean'
        sheet2['B1'] = 'Median'
        sheet2['C1'] = 'std'
        sheet2['A2'] = mean
        sheet2['B2'] = median
        sheet2['C2'] = std

        #Sheet 3 will be for the Barchart:
        #Creating a chart only for the top 10 values:
        refObj = openpyxl.chart.Reference( sheet, min_col=2, min_row=1,max_col =2, max_row =11)
        categories = openpyxl.chart.Reference( sheet, min_col=1, min_row=2, max_row =11)
        #seriesObj = openpyxl.chart.Series(refObj, title = 'Bitcoin Ranking')

        #Chossing the data
        chartObj = openpyxl.chart.BarChart()
        chartObj.add_data(refObj, titles_from_data=True)
        chartObj.set_categories(categories)


        chartObj.title = 'Bar Chart Rank vs Coin name'
        sheet3.add_chart(chartObj,'A1')
        wb.save('Test.xlsx') #"Excel/samplechart.xlsx"
        print('Done')

        #Sheet 4 will contain a Line Chart
        #Creating a line only for the top 10 values:
        refObj = openpyxl.chart.Reference( sheet, min_col=3, min_row=1,max_col =3, max_row =11)
        categories = openpyxl.chart.Reference( sheet, min_col=2, min_row=2, max_row =11)

        #Chossing the data
        chartObj = openpyxl.chart.LineChart()
        chartObj.add_data(refObj, titles_from_data=True)
        chartObj.set_categories(categories)

        chartObj.title = 'Line Chart Price vs Rank'
        sheet4.add_chart(chartObj,'A1')
        wb.save('static/files/Test.xlsx') #"Excel/samplechart.xlsx"
        print('Done')

        #Sheet 5 will contain a Pie Chart
        #Creating a line only for the top 10 values:

        sheet5['A1'] = 'Categories'
        sheet5['A2'] = 'Category 1'
        sheet5['A3'] = 'Category 2'
        sheet5['A4'] = 'Category 3'
        sheet5['B1'] = 'PERCENTAGE'
        sheet5['B2'] = 10
        sheet5['B3'] = 60
        sheet5['B4'] = 30

        #Chart data:      
        data = openpyxl.chart.Reference(sheet5, min_col=2, min_row=2, max_row=5)  # Values in the second column (Value)
        categories = openpyxl.chart.Reference(sheet5, min_col=1, min_row=2, max_row=5)  # Categories in the first column (Category)


        #Chossing the data
        chartObj = openpyxl.chart.PieChart()

        # Adding the information:
        chartObj.add_data(data, titles_from_data=True)
        chartObj.set_categories(categories)

        chartObj.title = 'Pie Chart'

        sheet5.add_chart(chartObj,'D1')
        
        # Set chart title and labels (optional)
        wb.save('Test.xlsx') #"Excel/samplechart.xlsx"
        print('Done')

        #Sheet 6 will contain a Scatter Plot:
        #Creating a scatter plot only for the top 10 values:
        values_x = openpyxl.chart.Reference( sheet, min_col=2, min_row=1,max_col =2, max_row =11)
        values_y = openpyxl.chart.Reference( sheet, min_col=3, min_row=1,max_col =3, max_row =11)
        Series_data = openpyxl.chart.Series( values_y, values_x, title = "Price vs Rank Scatterplot")

        #Chossing the data
        chartObj = openpyxl.chart.ScatterChart()
        chartObj.series.append(Series_data)

        chartObj.title = 'My Chart'
        #chartObj.append(seriesObj)
        sheet6.add_chart(chartObj,'A1')
        
        # Setting the title and x and y labels:
        chartObj.title = "Scatter Plot of Rank vs Price"
        chartObj.x_axis.title = "Rank"
        chartObj.y_axis.title = "Price in Usd"
        wb.save('Test.xlsx') #"Excel/samplechart.xlsx"
        print('Done')


        return
    except:
        return